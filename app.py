from flask import Flask, request, render_template, jsonify
import pandas as pd
import os
import json
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import io
import base64
from datetime import datetime
from collections import defaultdict
import openai
from dotenv import load_dotenv
import requests

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# OpenAI configuration
openai.api_key = os.getenv('OPENAI_API_KEY')  # Now loaded from .env file

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/parse', methods=['POST'])
def parse_data():
    try:
        data = request.get_json()
        file_data = data.get('file_data')
        filename = data.get('filename')
        
        if not file_data or not filename:
            return jsonify({'error': 'No file data provided'}), 400
        
        if not allowed_file(filename):
            return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx, .xls, or .xlsm)'}), 400
        
        # Decode the base64 file data
        try:
            file_bytes = base64.b64decode(file_data.split(',')[1])  # Remove data:... prefix
            file_obj = io.BytesIO(file_bytes)
        except Exception as e:
            return jsonify({'error': f'Error decoding file data: {str(e)}'}), 400
        
        # Read the Excel file and parse 'Table1' from 'Project Table' sheet
        try:
            # First, try to find the Table1 using openpyxl to get the exact range
            wb = load_workbook(file_obj, data_only=True)
            
            if 'Project Table' not in wb.sheetnames:
                return jsonify({'error': 'Sheet "Project Table" not found in the Excel file'}), 400
            
            project_sheet = wb['Project Table']
            
            # Look for Table1 in the worksheet tables
            table1_range = None
            table1_found = False
            
            # Check if there are any tables in the worksheet
            if hasattr(project_sheet, 'tables') and project_sheet.tables:
                print(f"Found {len(project_sheet.tables)} tables in worksheet")
                for table_name in project_sheet.tables:
                    print(f"Found table: {table_name}")
                    if table_name == 'Table1':
                        table_obj = project_sheet.tables[table_name]
                        table1_range = table_obj.ref
                        table1_found = True
                        print(f"Found Table1 with range: {table1_range}")
                        break
            else:
                print("No tables found in worksheet")
            
            if table1_found and table1_range:
                # Parse the range (e.g., "A1:D10")
                start_cell, end_cell = table1_range.split(':')
                
                # Reset file position for pandas
                file_obj.seek(0)
                
                # Use pandas to read just the table range
                # We'll read the sheet and then slice the specific range
                df_full = pd.read_excel(file_obj, sheet_name='Project Table', header=None)
                
                # Convert Excel range to row/column indices
                from openpyxl.utils import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(table1_range)
                
                # Extract just the table data (convert to 0-based indexing for pandas)
                table_data = df_full.iloc[min_row-1:max_row, min_col-1:max_col].copy()
                
                # Set the first row as column headers
                table_data.columns = table_data.iloc[0]
                table_data = table_data[1:].reset_index(drop=True)
                
            else:
                # Fallback: Look for "Table1" text in the sheet and read from there
                print("Table1 not found as named table, searching for 'Table1' text...")
                
                # Reset file position for pandas
                file_obj.seek(0)
                df = pd.read_excel(file_obj, sheet_name='Project Table')
                
                table_data = None
                # Search for "Table1" in the dataframe
                for i, row in df.iterroproject_sheet():
                    if any('Table1' in str(cell) for cell in row if pd.notna(cell)):
                        # Found Table1, start reading from the next row
                        table_data = df.iloc[i+1:].copy()
                        # Set first row as headers if they exist
                        if not table_data.empty:
                            table_data.columns = table_data.iloc[0]
                            table_data = table_data[1:].reset_index(drop=True)
                        break
                
                if table_data is None:
                    return jsonify({'error': 'Table1 not found in the Excel file. Please ensure there is a table named "Table1" or text "Table1" in the Project Table sheet.'}), 400
            
            # Clean the data
            # Remove any completely empty rows
            table_data = table_data.dropna(how='all')
            
            # Remove any completely empty columns
            table_data = table_data.dropna(axis=1, how='all')
            
            # Handle datetime columns - convert to string to avoid NaT serialization issues
            for col in table_data.columns:
                if table_data[col].dtype == 'datetime64[ns]':
                    # Convert datetime to string, handling NaT values
                    table_data[col] = table_data[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                    # Replace 'NaT' strings with None
                    table_data[col] = table_data[col].replace('NaT', None)
            
            # Convert to list of dictionaries (JSON format)
            # Replace NaN values with None for JSON serialization
            table_data = table_data.where(pd.notna(table_data), None)
            
            # Additional cleaning: replace any remaining NaN-like values
            table_data = table_data.replace([float('nan'), 'nan', 'NaN'], None)
            
            # Convert to list of dictionaries
            json_data = table_data.to_dict('records')
            
            # Final pass to ensure no NaN values in the JSON data
            import math
            for record in json_data:
                for key, value in record.items():
                    if isinstance(value, float) and math.isnan(value):
                        record[key] = None
            
            # Print to console as requested
            print("Parsed JSON data from Table1:")
            print(json.dumps(json_data, indent=2, default=str))
            
            # Now parse invoice data from "Invoice Data Imported" sheet
            invoice_data = []
            try:
                if 'Invoice Data Imported' in wb.sheetnames:
                    print("Found 'Invoice Data Imported' sheet, parsing invoice data...")
                    
                    # Reset file position for pandas
                    file_obj.seek(0)
                    
                    # Read the Invoice Data Imported sheet
                    invoice_df = pd.read_excel(file_obj, sheet_name='Invoice Data Imported')
                    
                    # Define the required columns
                    required_columns = ['invoice_id', 'project_code', 'invoice_date', 'payment_amount_usd', 'status']
                    
                    # Check which required columns exist in the sheet
                    available_columns = []
                    for col in required_columns:
                        # Check for exact match or case-insensitive match
                        matching_cols = [c for c in invoice_df.columns if c.lower() == col.lower()]
                        if matching_cols:
                            available_columns.append(matching_cols[0])  # Use the actual column name from Excel
                        else:
                            print(f"Warning: Column '{col}' not found in Invoice Data Imported sheet")
                    
                    if available_columns:
                        # Extract only the available required columns
                        invoice_subset = invoice_df[available_columns].copy()
                        
                        # Remove any completely empty rows
                        invoice_subset = invoice_subset.dropna(how='all')
                        
                        # Handle datetime columns - convert to string to avoid NaT serialization issues
                        for col in invoice_subset.columns:
                            if invoice_subset[col].dtype == 'datetime64[ns]':
                                # Convert datetime to string, handling NaT values
                                invoice_subset[col] = invoice_subset[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                                # Replace 'NaT' strings with None
                                invoice_subset[col] = invoice_subset[col].replace('NaT', None)
                        
                        # Convert to JSON format
                        invoice_subset = invoice_subset.where(pd.notna(invoice_subset), None)
                        
                        # Additional cleaning: replace any remaining NaN-like values
                        invoice_subset = invoice_subset.replace([float('nan'), 'nan', 'NaN'], None)
                        
                        invoice_data = invoice_subset.to_dict('records')
                        
                        # Final pass to ensure no NaN values in the invoice JSON data
                        import math
                        for record in invoice_data:
                            for key, value in record.items():
                                if isinstance(value, float) and math.isnan(value):
                                    record[key] = None
                        
                        print(f"Parsed {len(invoice_data)} invoice records")
                        print("Parsed Invoice Data:")
                        print(json.dumps(invoice_data[:3], indent=2, default=str))  # Show first 3 records
                        
                    else:
                        print("No required columns found in Invoice Data Imported sheet")
                        
                else:
                    print("Sheet 'Invoice Data Imported' not found in the Excel file")
                    
            except Exception as e:
                print(f"Error parsing invoice data: {str(e)}")
                # Continue with just project data if invoice parsing fails

            # Define month keys for 2024 and 2025
            month_keys_2024 = [f'2024 {month}' for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']]
            month_keys_2025 = [f'2025 {month}' for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']]
            all_month_keys = month_keys_2024 + month_keys_2025

            # Create month-to-date mapping for both years
            month_str_to_date = {
                **{f'2024 {month}': f'2024-{i:02d}-01 00:00:00' for i, month in enumerate(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], 1)},
                **{f'2025 {month}': f'2025-{i:02d}-01 00:00:00' for i, month in enumerate(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], 1)}
            }

            # Initialize totals dictionary
            monthly_totals = {key: 0 for key in all_month_keys}

            # Sum revenue for each month across all projects
            for record in json_data[:-1]:
                for key in all_month_keys:
                    value = record.get(key)
                    if isinstance(value, (int, float)):
                        monthly_totals[key] += value
                    elif isinstance(value, str):
                        try:
                            monthly_totals[key] += float(value.replace(',', '').strip())
                        except:
                            continue  # Skip malformed values

            # Create the rev_totals array
            rev_totals = [
                {
                    "revenue_total": monthly_totals[key],
                    "month": month_str_to_date[key]
                }
                for key in all_month_keys
            ]

            # Convert invoice date to datetime and amount to numeric
            for invoice in invoice_data:
                try:
                    invoice['invoice_date'] = pd.to_datetime(invoice.get('invoice_date'), errors='coerce')
                    invoice['payment_amount_usd'] = float(invoice.get('payment_amount_usd', 0))
                except:
                    invoice['payment_amount_usd'] = 0

            # Initialize monthly totals
            monthly_invoice_totals = defaultdict(float)

            # Aggregate per month
            for inv in invoice_data:
                inv_date = inv.get('invoice_date')
                amount = inv.get('payment_amount_usd', 0)
                if isinstance(inv_date, pd.Timestamp) and inv_date.year == 2025:
                    month_key = datetime(inv_date.year, inv_date.month, 1).strftime('%Y-%m-%d 00:00:00')
                    monthly_invoice_totals[month_key] += amount

            # Generate output list
            invoice_totals_2025 = [
                {"invoice_total": monthly_invoice_totals.get(f"2025-{month:02d}-01 00:00:00", 0.0),
                "month": f"2025-{month:02d}-01 00:00:00"}
                for month in range(1, 13)
            ]
            
            # Get current year dynamically and filter project data for current year
            current_year = datetime.now().year
            current_year_projects = [
                project for project in json_data 
                if project.get('Year') == current_year
            ]
            
            # Generate insights for current year projects using OpenAI
            project_insights = generate_project_insights(current_year_projects, current_year, rev_totals)
            
            return jsonify({
                'success': True,
                'project_data': json_data,
                'invoice_data': invoice_data,
                'rev_totals': rev_totals,
                'invoice_totals_2025': invoice_totals_2025,
                'current_year_projects': current_year_projects,
                'project_insights': project_insights,
                'message': f'Successfully parsed {len(json_data)} project rows and {len(invoice_data)} invoice records'
            })
            
        except ValueError as e:
            if "Worksheet named 'Project Table' not found" in str(e):
                return jsonify({'error': 'Sheet "Project Table" not found in the Excel file'}), 400
            else:
                return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 400
                
    except Exception as e:
        print(f"Error parsing data: {str(e)}")
        return jsonify({'error': f'Error parsing data: {str(e)}'}), 500

def calculate_current_year_revenue(rev_totals, current_year):
    """Calculate total revenue for current year from rev_totals array"""
    total_revenue = 0
    current_year_str = str(current_year)
    
    for month_data in rev_totals:
        month_date = month_data.get('month', '')
        if current_year_str in month_date:  # Check if the year is in the month date string
            total_revenue += month_data.get('revenue_total', 0)
    
    return total_revenue

def generate_project_insights(current_year_projects, current_year, rev_totals):
    """Generate insights using OpenAI API"""
    try:
        project_count = len(current_year_projects)
        total_revenue = calculate_current_year_revenue(rev_totals, current_year)
        avg_revenue = total_revenue / project_count if project_count > 0 else 0
        
        # Limit to first 10 projects for the prompt to avoid token limits
        sample_projects = current_year_projects[:10]
        
        # Custom datetime handler for JSON serialization
        def datetime_handler(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            elif hasattr(obj, 'isoformat'):  # pandas Timestamp
                return obj.isoformat()
            raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

        prompt = f"""
            ⚠️ 모든 분석 결과는 반드시 **한국어**로 작성해주세요. 영어는 숫자, 고유명사(회사명 등)에만 사용 가능합니다.

            {current_year}년의 {project_count}개 기술 프로젝트 데이터를 비즈니스 인사이트 관점에서 분석해주세요.

            🔹 핵심 지표 요약 (KEY METRICS SUMMARY):
            - 총 프로젝트 수: {project_count}
            - 총 수익: ${total_revenue:,.2f}
            - 프로젝트당 평균 수익: ${avg_revenue:,.2f}

            🔹 전체 프로젝트 데이터 (FULL PROJECT DATA):
            {json.dumps(sample_projects, indent=2, default=datetime_handler)}

            🔹 월별 수익 총합 데이터 (MONTHLY REVENUE TOTAL DATA):
            {json.dumps(rev_totals, indent=2)} 

            📌 다음 영역에 대해 종합적인 비즈니스 분석을 작성해주세요:

            1. # 경영 요약
            - 올해 프로젝트 포트폴리오 개요
            - 주요 성과 지표 및 핵심 수치 요약
            - 주요 인사이트 및 의미 있는 관찰 결과

            2. # 트렌드 분석
            - 월별 및 분기별 수익 패턴 및 차이 분석
            - 수익이 높은 시기와 낮은 시기 구분
            - 계절성 및 반복되는 경향 파악
            - 과거 데이터와의 비교 포함

            3. # 전략적 제언
            - 수익 패턴 기반 3~5개의 실행 가능한 전략 제안
            - 수익이 낮은 시기의 개선 전략 제시
            - AI 관련 최근 뉴스 및 트렌드를 반영한 제안 포함 (가능한 경우 출처 제시)
            - TecAce 고객사인 Samsung 및 SKT 관련 최근 뉴스 기반 전략 제안 (가능한 경우 출처 제시)
            - 전체 수익 개선을 위한 우선순위 영역 제안

            4. # 수익 목표 달성 확률 분석
            - 현재 프로젝트 데이터를 바탕으로 수익 달성 가능성 분석
            - 주요 리스크 요소 식별
            - 향후 수익 극대화를 위한 전략 제안

            5. # 수익 예측
            - 데이터 기반 향후 실적 예측
            - 수익 타깃에 대한 리스크 요인 파악
            - 대응 전략 제안

            6. # 리스크 평가
            - 수익 데이터 내 우려되는 패턴 및 경고 지점 파악
            - 리스크 대응 방안 제안
            - 집중 관리가 필요한 월 또는 분기 강조

            📝 보고서 형식:
            - # ## ### 제목 구분 사용
            - **굵은 글씨**로 핵심 강조
            - *기울임체*는 미묘한 강조에 사용
            - - 또는 * 로 목록 작성
            - 1. 2. 3. 순서 있는 목록 작성
            - 📊 📈 📉 ⚠️ 💡 🎯 이모지를 전략적으로 사용
            - 각 섹션 간에 충분한 여백을 두어 가독성 확보

            ✳️ **다시 한번 강조합니다: 모든 응답은 반드시 한국어로 작성해 주세요.**
                        🚨 **중요 알림**: 
            - 영어 단어나 문장은 절대 사용하지 마세요
            - 모든 제목, 분석, 설명은 한국어로만 작성
            - 회사명(Samsung, SKT 등)과 숫자만 원래 형태 유지
            - 한국 비즈니스 용어와 표현 사용 필수
            
            🇰🇷 **응답 언어: 한국어 100% 필수**
            """

#         prompt = f"""
# I have project data for {project_count} tech projects from {current_year} that I need you to analyze for business insights.
 
# KEY METRICS SUMMARY:
# - Total {current_year} projects: {project_count}
# - Total {current_year} revenue: ${total_revenue:,.2f}
# - Average revenue per {current_year} project: ${avg_revenue:,.2f}
 
# FULL PROJECT DATA:
# {json.dumps(sample_projects, indent=2, default=datetime_handler)}

# MONTHLY REVENUE TOTAL DATA:
# {json.dumps(rev_totals, indent=2)} 
 
# Please provide a comprehensive business analysis focusing on the following areas:
 
# 1. EXECUTIVE SUMMARY
#    - Overview of the current year's project portfolio's current state
#    - Key performance indicators and headline metrics
#    - High-level observations and critical insights
 
# 2. TREND ANALYSIS
#    - Identify monthly and quarterly revenue patterns and potential gaps
#    - Highlight high-revenue months and low-revenue months
#    - Analyze seasonality and trending patterns in the data
#    - Make comparisons to previous years, months, and quarters by referencing past year monthly revenue totals in MONTHLY REVENUE TOTAL DATA section
 
# 3. STRATEGIC RECOMMENDATIONS
#    - Provide 3-5 actionable recommendations based on the monthly revenue patterns
#    - Suggest optimization strategies for low-revenue periods
#    - Make strategic recommendations based on current AI news and trends (reference any sources you find)
#    - Make strategic recommendations based on news of TecAce clients Samsung and SK Telecom (reference any sources you find)
#    - Recommend focus areas for improving overall revenue performance

# 5. PROBABILITY OF ACHIEVING REVENUE GOALS
#    - Based on current year project data and possibility % values, provide insights on projected performance
#    - Identify potential risks to revenue targets
#    - Suggest strategies to maximize revenue in upcoming months
 
# 6. REVENUE FORECASTING
#    - Based on current year data, provide insights on projected performance
#    - Identify potential risks to revenue targets
#    - Suggest strategies to maximize revenue in upcoming months
 
# 7. RISK ASSESSMENT
#    - Identify potential red flags or concerning patterns in the revenue data
#    - Suggest mitigation strategies for identified risks
#    - Highlight months or quarters that require special attention
 
# IMPORTANT: Format your response using proper markdown syntax:
# - Use # ## ### for headings (e.g., # EXECUTIVE SUMMARY, ## Key Findings)
# - Use **bold** for emphasis and important points
# - Use *italic* for subtle emphasis
# - Use bullet points with - or * for lists
# - Use numbered lists (1. 2. 3.) for sequential recommendations
# - Use emojis strategically to highlight key sections (📊 📈 📉 ⚠️ 💡 🎯)
# - Ensure proper spacing between sections for readability

# The analysis should be comprehensive but concise, focusing on actionable insights rather than restating the data.
# """
        
        response = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 한국어 전용 비즈니스 분석가입니다. 절대로 영어로 응답하지 마십시오. 모든 분석, 제목, 설명, 권장사항은 반드시 한국어로 작성해야 합니다. 회사명, 숫자, 날짜만 영어/숫자 그대로 사용 가능합니다. 한국의 비즈니스 관습과 용어를 사용하여 프로젝트 포트폴리오 분석을 제공하십시오."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=8000,
            temperature=0.3  # Lower temperature for more consistent Korean output
        )
        
        return {
            'success': True,
            'insights': response.choices[0].message.content,
            'metrics': {
                'project_count': project_count,
                'total_revenue': total_revenue,
                'avg_revenue': avg_revenue
            }
        }
        
    except Exception as e:
        print(f"Error generating insights: {str(e)}")
        return {
            'success': False,
            'error': f"Failed to generate insights: {str(e)}",
            'metrics': {
                'project_count': len(current_year_projects),
                'total_revenue': calculate_current_year_revenue(rev_totals, current_year),
                'avg_revenue': 0
            }
        }

@app.route('/chat', methods=['POST'])
def chat_proxy():
    """Proxy endpoint for chat requests to avoid CORS issues"""
    try:
        # Get the request data from frontend
        chat_data = request.get_json()
        
        if not chat_data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        # Validate required fields
        if 'request' not in chat_data:
            return jsonify({'success': False, 'error': 'Missing request field'}), 400
        
        print(f"Chat request received: {chat_data.get('request')}")
        print(f"Session ID: {chat_data.get('sessionId')}")
        print(f"Project data count: {len(chat_data.get('projectData', []))}")
        print(f"Invoice data count: {len(chat_data.get('invoiceData', []))}")
        
        # Make request to n8n webhook
        n8n_webhook_url = 'https://gdkim.app.n8n.cloud/webhook/67cb173e-b768-43f2-99cc-45aaf49fc108'
        
        response = requests.post(
            n8n_webhook_url,
            json=chat_data,
            headers={'Content-Type': 'application/json'},
            timeout=60  # Increased to 60 second timeout
        )
        
        print(f"n8n response status: {response.status_code}")
        
        if response.status_code == 200:
            # Parse the JSON response from n8n and extract the "output" field
            try:
                response_json = response.json()
                print(f"Response type: {type(response_json)}")
                print(f"Response content preview: {str(response_json)[:200]}...")
                
                # Handle different response formats
                if isinstance(response_json, dict):
                    # If it's a dictionary, try to get 'output' field
                    output_text = response_json.get('output', response_json.get('response', response.text))
                elif isinstance(response_json, list):
                    # If it's a list, try to get the first item or convert to string
                    if len(response_json) > 0 and isinstance(response_json[0], dict):
                        output_text = response_json[0].get('output', response_json[0].get('response', str(response_json)))
                    else:
                        output_text = str(response_json)
                else:
                    # If it's something else, convert to string
                    output_text = str(response_json)
                
                return jsonify({
                    'success': True, 
                    'response': output_text
                })
            except json.JSONDecodeError:
                # Fallback to raw text if JSON parsing fails
                print("JSON decode error, using raw response text")
                return jsonify({
                    'success': True, 
                    'response': response.text
                })
        else:
            print(f"n8n error: {response.status_code} - {response.text}")
            return jsonify({
                'success': False, 
                'error': f'n8n webhook returned status {response.status_code}'
            }), response.status_code
            
    except requests.exceptions.Timeout:
        print("Request to n8n webhook timed out")
        return jsonify({
            'success': False, 
            'error': 'Request timeout - the AI service is processing your request but it\'s taking longer than expected (60+ seconds). Please try a simpler question or try again later.'
        }), 408
        
    except requests.exceptions.ConnectionError:
        print("Connection error to n8n webhook")
        return jsonify({
            'success': False, 
            'error': 'Connection error - unable to reach the AI service'
        }), 503
        
    except Exception as e:
        print(f"Unexpected error in chat proxy: {str(e)}")
        return jsonify({
            'success': False, 
            'error': 'An unexpected error occurred'
        }), 500

if __name__ == '__main__':
    # Get port from environment variable for deployment platforms like Render
    port = int(os.environ.get('PORT', 5000))
    # Set debug=False for production
    app.run(host='0.0.0.0', port=port, debug=False) 